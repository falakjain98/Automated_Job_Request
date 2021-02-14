#! python3
# sendCV.py - Send emails to potential recruiters using an excel sheet as status

# importing required modules
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl, smtplib, sys, time, pyautogui,smtplib

# mail body
mail_content = '''Hello,
Please find my CV attached with this mail.
My github: https://github.com/falakjain98
Thank You
'''
# Entering password
password = pyautogui.password('Please enter your password')

# Log in to email account
smtpObj = smtplib.SMTP('smtp-mail.outlook.com',587)
smtpObj.ehlo()
smtpObj.starttls()
try:
    smtpObj.login('jain.falak@hotmail.com',password)
except smtplib.SMTPAuthenticationError:
    print('Invalid password, run script again')
    sys.exit()

# Opening spreadhsset and getting dues status
wb = openpyxl.load_workbook('recruiters.xlsx')
sheet = wb['Sheet1']

# Getting current month using the time module
col_name = str(time.ctime()).split(' ')[1].split(' ')[0] + '_' + str(time.ctime())[-4:]

# Getting a list of all columns in the sheet
col_names = [Col[0].value for Col in sheet.iter_cols(1, sheet.max_column)]

# Checking if latest column exists in file and adding if it does not
if col_name not in col_names:
    last_col = sheet.max_column+1
    sheet.cell(row = 1,column = last_col).value = col_name
    print(f'Adding {col_name} to file')
else:
    last_col = sheet.max_column
    print(f'{col_name} already exists, proceeding...')

# Check each recruiter's sent status
unsent_recruiters = {}
for r in range(2,sheet.max_row+1):
    cv_sent = sheet.cell(row = r, column = last_col).value
    if cv_sent!='sent':
        name = sheet.cell(row = r, column = 1).value
        email = sheet.cell(row = r, column = 2).value
        company = sheet.cell(row = r, column = 3).value
        unsent_recruiters[name] = [email,company]
        sheet.cell(row = r,column = last_col).value = 'sent'

# Setup of MIME for pdf attachment
message = MIMEMultipart()
message['From'] = 'jain.falak@hotmail.com'
#message['To'] = ''
message['Subject'] = 'This is a test mail by Falak Jain.'
#The body and the attachments for the mail
message.attach(MIMEText(mail_content, 'plain'))
attach_file_name = 'Resume (Falak Jain).pdf'
attach_file = open(attach_file_name, 'rb') # Open the file as binary mode
payload = MIMEBase('application', 'octate-stream',Name = attach_file_name)
payload.set_payload((attach_file).read())
encoders.encode_base64(payload) #encode the attachment
#add payload header with filename
payload.add_header('Content-Decomposition', 'attachment', filename=attach_file_name)
message.attach(payload)
text = message.as_string()

# Send out CV emails
for name, details in unsent_recruiters.items():
    body = 'Subject: '
    print('Sending email to %s at %s' % (name,details[1]))
    message['To'] = details[0]
    #sendmailStatus = smtpObj.sendmail('jain.falak@hotmail.com',details[0],text)
    #if sendmailStatus != {}:
        #print('There was a problem sending an email to %s: %s' % (email,sendmailStatus))

# quitting email account
smtpObj.quit()

# saving modified recruiters.xlsx
#wb.save('recruiters.xlsx')